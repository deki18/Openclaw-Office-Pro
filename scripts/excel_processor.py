"""
Office Pro - Enterprise Document Automation Suite
Excel Processor Module

基于 openpyxl 和 xlsx-template 理念的企业级 Excel 处理
支持模板驱动、数据替换、图表生成
"""

from __future__ import annotations

import io
import json
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Tuple, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.workbook import Workbook as WorkbookType
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import (
        Font, Fill, PatternFill, Border, Side, Alignment, Protection,
        NamedStyle
    )
    from openpyxl.utils import get_column_letter, coordinate_to_tuple
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, ScatterChart, Reference,
        Series
    )
    from openpyxl.chart.label import DataLabelList
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.comments import Comment
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class XlsxTemplateEngine:
    """
    xlsx-template 风格的 Excel 模板引擎
    
    支持在 Excel 模板中使用占位符，然后替换为实际数据
    占位符格式：${variable}、${table:data}、${image:logo}
    """
    
    # 占位符正则表达式
    PLACEHOLDER_PATTERN = re.compile(r'\$\{([^}]+)\}')
    
    def __init__(self, workbook: WorkbookType):
        """
        初始化模板引擎
        
        Args:
            workbook: openpyxl Workbook 对象
        """
        self.workbook = workbook
        self.substitutions: Dict[str, Any] = {}
    
    def substitute(self, data: Dict[str, Any]) -> None:
        """
        执行数据替换
        
        Args:
            data: 替换数据字典
        """
        self.substitutions = data
        
        # 遍历所有工作表
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            self._process_worksheet(worksheet)
    
    def _process_worksheet(self, worksheet: Worksheet) -> None:
        """处理工作表中的所有单元格"""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    new_value = self._replace_placeholders(cell.value)
                    if new_value != cell.value:
                        cell.value = new_value
    
    def _replace_placeholders(self, text: str) -> Any:
        """
        替换文本中的占位符
        
        支持的格式：
        - ${variable} - 简单变量替换
        - ${table:data.property} - 表格数据
        - ${image:logo} - 图片插入（返回特殊标记）
        """
        def replace_match(match):
            placeholder = match.group(1).strip()
            
            # 处理 table: 前缀
            if placeholder.startswith('table:'):
                return self._handle_table_placeholder(placeholder[6:])
            
            # 处理 image: 前缀
            if placeholder.startswith('image:'):
                return self._handle_image_placeholder(placeholder[6:])
            
            # 简单变量替换
            if placeholder in self.substitutions:
                value = self.substitutions[placeholder]
                return self._format_value(value)
            
            # 尝试解析点号路径（如 user.name）
            if '.' in placeholder:
                value = self._get_nested_value(self.substitutions, placeholder)
                if value is not None:
                    return self._format_value(value)
            
            # 未找到替换值，保留原样
            return match.group(0)
        
        return self.PLACEHOLDER_PATTERN.sub(replace_match, text)
    
    def _handle_table_placeholder(self, path: str) -> str:
        """处理表格占位符"""
        # 这里应该返回表格数据，但在单元格中需要特殊处理
        # 简化处理：尝试获取值并格式化
        value = self._get_nested_value(self.substitutions, path)
        return self._format_value(value) if value is not None else f"${{table:{path}}}"
    
    def _handle_image_placeholder(self, path: str) -> str:
        """处理图片占位符"""
        # 图片需要特殊处理，这里返回标记
        return f"${{image:{path}}}"
    
    def _get_nested_value(self, data: Dict, path: str) -> Any:
        """获取嵌套字典值"""
        keys = path.split('.')
        value = data
        for key in keys:
            if isinstance(value, dict) and key in value:
                value = value[key]
            else:
                return None
        return value
    
    def _format_value(self, value: Any) -> str:
        """格式化值为字符串"""
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")
        return str(value)


class ExcelProcessor:
    """
    企业级 Excel 处理器
    
    支持功能：
    - 工作簿创建与编辑
    - 模板驱动数据替换（xlsx-template 风格）
    - 图表生成
    - 数据透视表
    - 格式设置
    """
    
    def __init__(self, template_dir: Optional[str] = None):
        """
        初始化 Excel 处理器
        
        Args:
            template_dir: 模板目录路径
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required. Install with: pip install openpyxl"
            )
        
        self.template_dir = template_dir or self._get_default_template_dir()
        self._workbook: Optional[WorkbookType] = None
        self._template_engine: Optional[XlsxTemplateEngine] = None
    
    def _get_default_template_dir(self) -> str:
        """获取默认模板目录"""
        skill_root = Path(__file__).parent.parent
        templates_dir = skill_root / "assets" / "templates" / "excel"
        return str(templates_dir)
    
    # ==================== 工作簿操作 ====================
    
    def create_workbook(self) -> WorkbookType:
        """
        创建新工作簿
        
        Returns:
            Workbook 对象
        """
        self._workbook = Workbook()
        self._template_engine = None
        return self._workbook
    
    def load_workbook(self, path: str, data_only: bool = False) -> WorkbookType:
        """
        加载现有工作簿
        
        Args:
            path: 文件路径
            data_only: 是否只读取数据（不读取公式）
            
        Returns:
            Workbook 对象
        """
        self._workbook = load_workbook(path, data_only=data_only)
        self._template_engine = None
        return self._workbook
    
    def load_template(self, template_name: str) -> WorkbookType:
        """
        加载模板文件并初始化模板引擎
        
        Args:
            template_name: 模板文件名
            
        Returns:
            Workbook 对象
        """
        template_path = Path(self.template_dir) / template_name
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        self.load_workbook(str(template_path))
        self._template_engine = XlsxTemplateEngine(self._workbook)
        return self._workbook
    
    def save(self, path: str) -> None:
        """
        保存工作簿
        
        Args:
            path: 保存路径
        """
        if not self._workbook:
            raise RuntimeError("No workbook to save.")
        
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        self._workbook.save(path)
    
    # ==================== 模板渲染 ====================
    
    def render_template(self, data: Dict[str, Any]) -> WorkbookType:
        """
        使用数据渲染模板
        
        Args:
            data: 替换数据字典
            
        Returns:
            渲染后的 Workbook 对象
        """
        if not self._template_engine:
            raise RuntimeError("No template loaded. Call load_template() first.")
        
        self._template_engine.substitute(data)
        return self._workbook
    
    def render_and_save(self, data: Dict[str, Any], output_path: str) -> str:
        """
        渲染模板并保存
        
        Args:
            data: 替换数据字典
            output_path: 输出路径
            
        Returns:
            保存的文件路径
        """
        self.render_template(data)
        self.save(output_path)
        return output_path
    
    # ==================== 工作表操作 ====================
    
    def get_sheet(self, name: Optional[str] = None) -> Worksheet:
        """
        获取工作表
        
        Args:
            name: 工作表名称，默认获取活动工作表
            
        Returns:
            Worksheet 对象
        """
        if not self._workbook:
            raise RuntimeError("No workbook loaded.")
        
        if name:
            return self._workbook[name]
        return self._workbook.active
    
    def create_sheet(self, title: str, index: Optional[int] = None) -> Worksheet:
        """
        创建工作表
        
        Args:
            title: 工作表标题
            index: 插入位置
            
        Returns:
            Worksheet 对象
        """
        if not self._workbook:
            raise RuntimeError("No workbook loaded.")
        
        return self._workbook.create_sheet(title=title, index=index)
    
    def remove_sheet(self, name: str) -> None:
        """
        删除工作表
        
        Args:
            name: 工作表名称
        """
        if not self._workbook:
            raise RuntimeError("No workbook loaded.")
        
        sheet = self._workbook[name]
        self._workbook.remove(sheet)
    
    # ==================== 单元格操作 ====================
    
    def write_cell(self, cell: str, value: Any, sheet: Optional[str] = None) -> None:
        """
        写入单元格
        
        Args:
            cell: 单元格坐标（如 'A1'）
            value: 值
            sheet: 工作表名称
        """
        ws = self.get_sheet(sheet)
        ws[cell] = value
    
    def read_cell(self, cell: str, sheet: Optional[str] = None) -> Any:
        """
        读取单元格
        
        Args:
            cell: 单元格坐标
            sheet: 工作表名称
            
        Returns:
            单元格值
        """
        ws = self.get_sheet(sheet)
        return ws[cell].value
    
    def write_range(self, start_cell: str, data: List[List[Any]], sheet: Optional[str] = None) -> None:
        """
        写入数据区域
        
        Args:
            start_cell: 起始单元格
            data: 二维数据列表
            sheet: 工作表名称
        """
        ws = self.get_sheet(sheet)
        
        # 解析起始单元格
        from openpyxl.utils import coordinate_to_tuple
        start_row, start_col = coordinate_to_tuple(start_cell)
        
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(
                    row=start_row + row_idx,
                    column=start_col + col_idx,
                    value=value
                )
    
    # ==================== 样式与格式 ====================
    
    def set_cell_style(self, cell: str, 
                       font: Optional[Dict] = None,
                       fill: Optional[Dict] = None,
                       border: Optional[Dict] = None,
                       alignment: Optional[Dict] = None,
                       number_format: Optional[str] = None,
                       sheet: Optional[str] = None) -> None:
        """
        设置单元格样式
        
        Args:
            cell: 单元格坐标
            font: 字体设置 {'name': 'Arial', 'size': 12, 'bold': True}
            fill: 填充设置 {'color': 'FFFF00', 'pattern': 'solid'}
            border: 边框设置 {'style': 'thin', 'color': '000000'}
            alignment: 对齐设置 {'horizontal': 'center', 'vertical': 'center'}
            number_format: 数字格式 '#,##0.00'
            sheet: 工作表名称
        """
        ws = self.get_sheet(sheet)
        cell_obj = ws[cell]
        
        if font:
            cell_obj.font = Font(**font)
        if fill:
            cell_obj.fill = PatternFill(**fill)
        if border:
            side = Side(**border)
            cell_obj.border = Border(left=side, right=side, top=side, bottom=side)
        if alignment:
            cell_obj.alignment = Alignment(**alignment)
        if number_format:
            cell_obj.number_format = number_format
    
    def set_column_width(self, column: str, width: float, sheet: Optional[str] = None) -> None:
        """
        设置列宽
        
        Args:
            column: 列字母（如 'A'）
            width: 宽度
            sheet: 工作表名称
        """
        ws = self.get_sheet(sheet)
        ws.column_dimensions[column].width = width
    
    def set_row_height(self, row: int, height: float, sheet: Optional[str] = None) -> None:
        """
        设置行高
        
        Args:
            row: 行号（1-based）
            height: 高度
            sheet: 工作表名称
        """
        ws = self.get_sheet(sheet)
        ws.row_dimensions[row].height = height
    
    def merge_cells(self, range_str: str, sheet: Optional[str] = None) -> None:
        """
        合并单元格
        
        Args:
            range_str: 区域（如 'A1:B2'）
            sheet: 工作表名称
        """
        ws = self.get_sheet(sheet)
        ws.merge_cells(range_str)
    
    # ==================== 图表 ====================
    
    def add_chart(self, chart_type: str, data_range: str, 
                  title: Optional[str] = None,
                  position: Optional[str] = None,
                  sheet: Optional[str] = None) -> Any:
        """
        添加图表
        
        Args:
            chart_type: 图表类型（bar/line/pie/scatter）
            data_range: 数据区域
            title: 图表标题
            position: 图表位置（单元格坐标）
            sheet: 工作表名称
            
        Returns:
            Chart 对象
        """
        ws = self.get_sheet(sheet)
        
        # 创建图表
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        elif chart_type == 'scatter':
            chart = ScatterChart()
        else:
            chart = BarChart()
        
        # 设置数据
        chart.add_data(Reference(ws, data_range))
        
        # 设置标题
        if title:
            chart.title = title
        
        # 设置位置
        if position:
            ws.add_chart(chart, position)
        else:
            ws.add_chart(chart, 'E5')
        
        return chart
    
    # ==================== 数据透视表 ====================
    
    def create_pivot_table(self, source_range: str, dest_cell: str,
                          rows: Optional[List[str]] = None,
                          columns: Optional[List[str]] = None,
                          values: Optional[List[Tuple[str, str]]] = None,
                          sheet: Optional[str] = None) -> Any:
        """
        创建数据透视表（简化版）
        
        注意：openpyxl 的数据透视表功能有限，建议使用 pandas 预处理数据
        
        Args:
            source_range: 源数据区域
            dest_cell: 目标位置
            rows: 行字段列表
            columns: 列字段列表
            values: 值字段列表 [(字段名, 聚合函数)]
            sheet: 工作表名称
        """
        # 数据透视表实现较复杂，建议使用 pandas 处理后再写入
        # 这里返回提示信息
        raise NotImplementedError(
            "数据透视表建议使用 pandas 预处理数据后再写入 Excel。"
            "示例：pivot_df = df.pivot_table(...) ; ep.write_dataframe(pivot_df, ...)"
        )
    
    # ==================== 数据导入导出 ====================
    
    def read_dataframe(self, sheet: Optional[str] = None, 
                       header: int = 0, 
                       range_str: Optional[str] = None) -> Any:
        """
        读取数据到 pandas DataFrame
        
        Args:
            sheet: 工作表名称
            header: 表头行号
            range_str: 数据区域
            
        Returns:
            pandas DataFrame
        """
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas is required. Install with: pip install pandas")
        
        ws = self.get_sheet(sheet)
        
        # 读取数据
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        if not data:
            return pd.DataFrame()
        
        # 创建 DataFrame
        df = pd.DataFrame(data[header+1:], columns=data[header])
        return df
    
    def write_dataframe(self, df: Any, start_cell: str = 'A1', 
                        sheet: Optional[str] = None,
                        include_header: bool = True,
                        index: bool = False) -> None:
        """
        将 pandas DataFrame 写入 Excel
        
        Args:
            df: pandas DataFrame
            start_cell: 起始单元格
            sheet: 工作表名称
            include_header: 是否包含表头
            index: 是否包含索引
        """
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas is required. Install with: pip install pandas")
        
        ws = self.get_sheet(sheet)
        
        # 解析起始单元格
        from openpyxl.utils import coordinate_to_tuple
        start_row, start_col = coordinate_to_tuple(start_cell)
        
        # 写入索引（如果需要）
        if index:
            for i, idx in enumerate(df.index):
                ws.cell(row=start_row + i + (1 if include_header else 0), 
                       column=start_col, value=idx)
            start_col += 1
        
        # 写入表头
        if include_header:
            for col_idx, col_name in enumerate(df.columns):
                ws.cell(row=start_row, column=start_col + col_idx, value=col_name)
            start_row += 1
        
        # 写入数据
        for row_idx, row in enumerate(df.itertuples(index=False)):
            for col_idx, value in enumerate(row):
                ws.cell(row=start_row + row_idx, 
                       column=start_col + col_idx, 
                       value=value)
    
    def import_csv(self, csv_path: str, sheet: Optional[str] = None, 
                   delimiter: str = ',',
                   encoding: str = 'utf-8') -> None:
        """
        从 CSV 文件导入数据
        
        Args:
            csv_path: CSV 文件路径
            sheet: 目标工作表名称
            delimiter: 分隔符
            encoding: 编码
        """
        import csv
        
        ws = self.get_sheet(sheet)
        
        with open(csv_path, 'r', encoding=encoding) as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
    
    def export_csv(self, csv_path: str, sheet: Optional[str] = None,
                   delimiter: str = ',',
                   encoding: str = 'utf-8') -> None:
        """
        导出到 CSV 文件
        
        Args:
            csv_path: 输出 CSV 文件路径
            sheet: 源工作表名称
            delimiter: 分隔符
            encoding: 编码
        """
        import csv
        
        ws = self.get_sheet(sheet)
        
        with open(csv_path, 'w', encoding=encoding, newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
    
    # ==================== 属性访问 ====================
    
    @property
    def workbook(self) -> Optional[WorkbookType]:
        """获取当前工作簿"""
        return self._workbook
    
    @property
    def sheetnames(self) -> List[str]:
        """获取所有工作表名称"""
        if not self._workbook:
            return []
        return self._workbook.sheetnames
